"""
Servicio de reportes y exportacion.
Sprint 7.2 - Generacion de reportes en Excel, CSV y PDF.

Funcionalidades:
- Exportar solicitudes a Excel/CSV
- Exportar inventario y alertas MRP
- Generar reportes de KPIs
- Reportes personalizados con filtros
"""

import csv
import logging
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def get_db_connection():
    """Obtiene conexion a base de datos."""
    try:
        from backend.core.db import get_db_connection as db_conn

        return db_conn()
    except ImportError:
        from core.db import get_db_connection as db_conn

        return db_conn()


class ReportingService:
    """
    Servicio para generacion de reportes y exportacion de datos.

    Formatos soportados:
    - xlsx: Excel (requiere openpyxl)
    - csv: Comma-Separated Values
    - pdf: PDF basico (requiere reportlab)
    """

    SUPPORTED_FORMATS = ["xlsx", "csv", "pdf"]

    def __init__(self):
        """Inicializa el servicio de reportes."""
        self._check_dependencies()

    def _check_dependencies(self):
        """Verifica dependencias disponibles."""
        self._has_openpyxl = False
        self._has_reportlab = False

        try:
            import openpyxl

            self._has_openpyxl = True
        except ImportError:
            logger.warning("openpyxl no disponible, Excel export limitado")

        try:
            import reportlab

            self._has_reportlab = True
        except ImportError:
            logger.warning("reportlab no disponible, PDF export no disponible")

    def get_supported_formats(self) -> List[str]:
        """Retorna formatos soportados."""
        return self.SUPPORTED_FORMATS.copy()

    # ========================================================================
    # EXPORTAR SOLICITUDES
    # ========================================================================

    def export_solicitudes(
        self,
        solicitudes: List[Dict[str, Any]],
        formato: str = "xlsx",
        filtros: Optional[Dict[str, Any]] = None,
        columnas: Optional[List[str]] = None,
        incluir_metadatos: bool = False,
    ) -> Dict[str, Any]:
        """
        Exporta lista de solicitudes al formato especificado.

        Args:
            solicitudes: Lista de solicitudes
            formato: Formato de salida (xlsx, csv, pdf)
            filtros: Filtros aplicados (para metadata)
            columnas: Columnas a incluir (None = todas)
            incluir_metadatos: Incluir info de generacion

        Returns:
            Dict con contenido, filename y metadata
        """
        if formato not in self.SUPPORTED_FORMATS:
            return {
                "success": False,
                "error": f"Formato no soportado: {formato}. Use: {self.SUPPORTED_FORMATS}",
            }

        try:
            # Filtrar columnas si se especifican
            if columnas and solicitudes:
                solicitudes = [{k: v for k, v in s.items() if k in columnas} for s in solicitudes]

            # Generar contenido segun formato
            if formato == "xlsx":
                contenido = self._generate_excel(solicitudes, "Solicitudes")
            elif formato == "csv":
                contenido = self._generate_csv(solicitudes)
            else:  # pdf
                contenido = self._generate_pdf_simple(solicitudes, "Reporte de Solicitudes")

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"solicitudes_{timestamp}.{formato}"

            result = {
                "success": True,
                "formato": formato,
                "contenido": contenido,
                "filename": filename,
                "total_registros": len(solicitudes),
                "generado_en": datetime.now(timezone.utc).isoformat(),
            }

            if filtros:
                result["filtros_aplicados"] = filtros

            if incluir_metadatos:
                result["metadatos"] = {
                    "generado_en": result["generado_en"],
                    "total_registros": len(solicitudes),
                    "formato": formato,
                }
                result["generado_por"] = "ReportingService"

            return result

        except Exception as e:
            logger.error(f"Error exportando solicitudes: {e}")
            return {"success": False, "error": str(e)}

    def export_solicitudes_from_db(
        self, formato: str = "xlsx", filtros: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Exporta solicitudes directamente desde BD.

        Args:
            formato: Formato de salida
            filtros: Filtros a aplicar (estado, centro, fecha, etc.)

        Returns:
            Dict con contenido exportado
        """
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                query = """
                    SELECT id, codigo, estado, criticidad, total_monto,
                           created_at, solicitante_id, centro
                    FROM solicitudes
                    WHERE 1=1
                """
                params = []

                if filtros:
                    if "estado" in filtros:
                        query += " AND estado = ?"
                        params.append(filtros["estado"])
                    if "centro" in filtros:
                        query += " AND centro = ?"
                        params.append(filtros["centro"])
                    if "fecha_desde" in filtros:
                        query += " AND created_at >= ?"
                        params.append(filtros["fecha_desde"])
                    if "fecha_hasta" in filtros:
                        query += " AND created_at <= ?"
                        params.append(filtros["fecha_hasta"])

                query += " ORDER BY created_at DESC"

                cursor.execute(query, params)
                solicitudes = [dict(row) for row in cursor.fetchall()]

            return self.export_solicitudes(
                solicitudes=solicitudes, formato=formato, filtros=filtros
            )

        except Exception as e:
            logger.error(f"Error exportando desde BD: {e}")
            return {"success": False, "error": str(e)}

    # ========================================================================
    # EXPORTAR INVENTARIO
    # ========================================================================

    def export_inventario(
        self, materiales: List[Dict[str, Any]], formato: str = "xlsx", incluir_alertas: bool = False
    ) -> Dict[str, Any]:
        """
        Exporta inventario de materiales.

        Args:
            materiales: Lista de materiales
            formato: Formato de salida
            incluir_alertas: Marcar materiales con alertas

        Returns:
            Dict con contenido exportado
        """
        if formato not in self.SUPPORTED_FORMATS:
            return {"success": False, "error": f"Formato no soportado: {formato}"}

        try:
            materiales_criticos = 0

            # Analizar alertas si se solicita
            if incluir_alertas:
                for mat in materiales:
                    stock = mat.get("stock_actual", 0) or 0
                    stock_seg = mat.get("stock_seguridad", 0) or 0
                    punto_pedido = mat.get("punto_pedido", 0) or 0

                    if stock < stock_seg:
                        mat["_alerta"] = "CRITICO"
                        materiales_criticos += 1
                    elif stock < punto_pedido:
                        mat["_alerta"] = "BAJO"
                    else:
                        mat["_alerta"] = "OK"

            # Generar contenido
            if formato == "xlsx":
                contenido = self._generate_excel(materiales, "Inventario")
            elif formato == "csv":
                contenido = self._generate_csv(materiales)
            else:
                contenido = self._generate_pdf_simple(materiales, "Inventario")

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

            return {
                "success": True,
                "formato": formato,
                "contenido": contenido,
                "filename": f"inventario_{timestamp}.{formato}",
                "total_registros": len(materiales),
                "materiales_criticos": materiales_criticos if incluir_alertas else None,
                "generado_en": datetime.now(timezone.utc).isoformat(),
            }

        except Exception as e:
            logger.error(f"Error exportando inventario: {e}")
            return {"success": False, "error": str(e)}

    def export_inventario_from_db(
        self, formato: str = "xlsx", centro: Optional[str] = None
    ) -> Dict[str, Any]:
        """Exporta inventario desde BD."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                query = """
                    SELECT codigo, descripcion, stock_actual, stock_seguridad,
                           punto_pedido, precio_usd, centro
                    FROM catalogo_materiales
                    WHERE activo = 1
                """
                params = []

                if centro:
                    query += " AND centro = ?"
                    params.append(centro)

                cursor.execute(query, params)
                materiales = [dict(row) for row in cursor.fetchall()]

            return self.export_inventario(
                materiales=materiales, formato=formato, incluir_alertas=True
            )

        except Exception as e:
            logger.error(f"Error exportando inventario desde BD: {e}")
            return {"success": False, "error": str(e)}

    # ========================================================================
    # EXPORTAR ALERTAS MRP
    # ========================================================================

    def export_alertas_mrp(
        self, alertas: List[Dict[str, Any]], formato: str = "xlsx"
    ) -> Dict[str, Any]:
        """
        Exporta alertas MRP.

        Args:
            alertas: Lista de alertas
            formato: Formato de salida

        Returns:
            Dict con contenido exportado
        """
        if formato not in self.SUPPORTED_FORMATS:
            return {"success": False, "error": f"Formato no soportado: {formato}"}

        try:
            if formato == "xlsx":
                contenido = self._generate_excel(alertas, "Alertas MRP")
            elif formato == "csv":
                contenido = self._generate_csv(alertas)
            else:
                contenido = self._generate_pdf_simple(alertas, "Alertas MRP")

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

            return {
                "success": True,
                "formato": formato,
                "contenido": contenido,
                "filename": f"alertas_mrp_{timestamp}.{formato}",
                "total_alertas": len(alertas),
                "generado_en": datetime.now(timezone.utc).isoformat(),
            }

        except Exception as e:
            logger.error(f"Error exportando alertas: {e}")
            return {"success": False, "error": str(e)}

    # ========================================================================
    # REPORTE DE KPIs
    # ========================================================================

    def generate_kpi_report(
        self,
        kpis: Dict[str, Any],
        formato: str = "xlsx",
        periodo_inicio: Optional[str] = None,
        periodo_fin: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Genera reporte de KPIs.

        Args:
            kpis: Diccionario de KPIs
            formato: Formato de salida
            periodo_inicio: Inicio del periodo
            periodo_fin: Fin del periodo

        Returns:
            Dict con contenido del reporte
        """
        if formato not in self.SUPPORTED_FORMATS:
            return {"success": False, "error": f"Formato no soportado: {formato}"}

        try:
            # Convertir KPIs a lista para exportar
            kpi_list = [{"metrica": k, "valor": v} for k, v in kpis.items()]

            if formato == "xlsx":
                contenido = self._generate_excel(kpi_list, "KPIs")
            elif formato == "csv":
                contenido = self._generate_csv(kpi_list)
            else:
                contenido = self._generate_pdf_simple(kpi_list, "Reporte de KPIs")

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

            result = {
                "success": True,
                "formato": formato,
                "contenido": contenido,
                "filename": f"kpis_{timestamp}.{formato}",
                "generado_en": datetime.now(timezone.utc).isoformat(),
            }

            if periodo_inicio or periodo_fin:
                result["periodo"] = {"inicio": periodo_inicio, "fin": periodo_fin}

            return result

        except Exception as e:
            logger.error(f"Error generando reporte KPIs: {e}")
            return {"success": False, "error": str(e)}

    # ========================================================================
    # REPORTE PERSONALIZADO
    # ========================================================================

    def generate_custom_report(
        self,
        titulo: str,
        datos: List[Dict[str, Any]],
        columnas: List[str],
        formato: str = "xlsx",
        agrupar_por: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Genera reporte personalizado.

        Args:
            titulo: Titulo del reporte
            datos: Datos a incluir
            columnas: Columnas a incluir
            formato: Formato de salida
            agrupar_por: Campo para agrupar datos

        Returns:
            Dict con contenido del reporte
        """
        if formato not in self.SUPPORTED_FORMATS:
            return {"success": False, "error": f"Formato no soportado: {formato}"}

        try:
            # Filtrar columnas
            datos_filtrados = [{k: v for k, v in d.items() if k in columnas} for d in datos]

            # Agrupar si se solicita
            if agrupar_por and agrupar_por in columnas:
                datos_filtrados = sorted(datos_filtrados, key=lambda x: x.get(agrupar_por, ""))

            if formato == "xlsx":
                contenido = self._generate_excel(datos_filtrados, titulo)
            elif formato == "csv":
                contenido = self._generate_csv(datos_filtrados)
            else:
                contenido = self._generate_pdf_simple(datos_filtrados, titulo)

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            safe_titulo = titulo.lower().replace(" ", "_")[:30]

            return {
                "success": True,
                "titulo": titulo,
                "formato": formato,
                "contenido": contenido,
                "filename": f"{safe_titulo}_{timestamp}.{formato}",
                "total_registros": len(datos),
                "generado_en": datetime.now(timezone.utc).isoformat(),
            }

        except Exception as e:
            logger.error(f"Error generando reporte personalizado: {e}")
            return {"success": False, "error": str(e)}

    # ========================================================================
    # GENERADORES DE FORMATO
    # ========================================================================

    def _generate_excel(self, datos: List[Dict[str, Any]], sheet_name: str = "Datos") -> bytes:
        """
        Genera archivo Excel.

        Args:
            datos: Lista de diccionarios
            sheet_name: Nombre de la hoja

        Returns:
            Bytes del archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Max 31 chars

            if not datos:
                ws.append(["Sin datos"])
            else:
                # Headers
                headers = list(datos[0].keys())
                ws.append(headers)

                # Estilo headers
                header_font = Font(bold=True)
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                for col, _ in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Datos
                for row_data in datos:
                    ws.append([row_data.get(h, "") for h in headers])

                # Auto-ajustar ancho columnas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

            # Guardar a bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except ImportError:
            # Fallback a CSV si openpyxl no esta disponible
            logger.warning("openpyxl no disponible, generando CSV")
            return self._generate_csv(datos)

    def _generate_csv(self, datos: List[Dict[str, Any]]) -> bytes:
        """
        Genera archivo CSV.

        Args:
            datos: Lista de diccionarios

        Returns:
            Bytes del archivo CSV
        """
        output = StringIO()

        if not datos:
            output.write("Sin datos\n")
        else:
            headers = list(datos[0].keys())
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerows(datos)

        return output.getvalue().encode("utf-8")

    def _generate_pdf_simple(self, datos: List[Dict[str, Any]], titulo: str) -> bytes:
        """
        Genera PDF simple (texto).

        Args:
            datos: Lista de diccionarios
            titulo: Titulo del documento

        Returns:
            Bytes del archivo PDF
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            output = BytesIO()
            c = canvas.Canvas(output, pagesize=letter)
            width, height = letter

            # Titulo
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, height - 50, titulo)

            # Datos
            c.setFont("Helvetica", 10)
            y = height - 80

            if not datos:
                c.drawString(50, y, "Sin datos")
            else:
                for i, row in enumerate(datos[:50]):  # Limitar a 50 filas
                    text = ", ".join(f"{k}: {v}" for k, v in row.items())[:100]
                    c.drawString(50, y, text)
                    y -= 15
                    if y < 50:
                        c.showPage()
                        y = height - 50

            c.save()
            output.seek(0)
            return output.getvalue()

        except ImportError:
            # Fallback a texto plano
            logger.warning("reportlab no disponible, generando texto plano")
            text = f"{titulo}\n\n"
            for row in datos:
                text += str(row) + "\n"
            return text.encode("utf-8")


# Instancia singleton
_reporting_service_instance: Optional[ReportingService] = None


def get_reporting_service() -> ReportingService:
    """Obtiene instancia singleton del servicio de reportes."""
    global _reporting_service_instance
    if _reporting_service_instance is None:
        _reporting_service_instance = ReportingService()
    return _reporting_service_instance
