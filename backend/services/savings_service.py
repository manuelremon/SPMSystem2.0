"""
Servicio de tracking de ahorros de costos.
Proporciona funciones para registrar ahorros, definir metas y generar reportes.
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from backend.core.db import get_db_connection, get_db_transaction, is_using_postgresql


def registrar_ahorro(data: Dict[str, Any], user_id: int) -> int:
    """
    Registra un ahorro de costo.

    Args:
        data: Diccionario con los campos del ahorro
        user_id: ID del usuario que registra el ahorro

    Returns:
        ID del ahorro registrado
    """
    with get_db_transaction() as (conn, cursor):
        if is_using_postgresql():
            cursor.execute("""
                INSERT INTO ahorro_costo (
                    categoria, tipo, monto, moneda, material_codigo,
                    proveedor_cuit, solicitud_id, descripcion,
                    registrado_por, fecha_realizacion
                ) VALUES (
                    %(categoria)s, %(tipo)s, %(monto)s, %(moneda)s, %(material_codigo)s,
                    %(proveedor_cuit)s, %(solicitud_id)s, %(descripcion)s,
                    %(registrado_por)s, %(fecha_realizacion)s
                )
                RETURNING id
            """, {
                'categoria': data['categoria'],
                'tipo': data['tipo'],
                'monto': data['monto'],
                'moneda': data.get('moneda', 'ARS'),
                'material_codigo': data.get('material_codigo'),
                'proveedor_cuit': data.get('proveedor_cuit'),
                'solicitud_id': data.get('solicitud_id'),
                'descripcion': data.get('descripcion'),
                'registrado_por': user_id,
                'fecha_realizacion': data.get('fecha_realizacion', datetime.now().isoformat())
            })
            ahorro_id = cursor.fetchone()[0]
        else:
            cursor.execute("""
                INSERT INTO ahorro_costo (
                    categoria, tipo, monto, moneda, material_codigo,
                    proveedor_cuit, solicitud_id, descripcion,
                    registrado_por, fecha_realizacion
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data['categoria'],
                data['tipo'],
                data['monto'],
                data.get('moneda', 'ARS'),
                data.get('material_codigo'),
                data.get('proveedor_cuit'),
                data.get('solicitud_id'),
                data.get('descripcion'),
                user_id,
                data.get('fecha_realizacion', datetime.now().isoformat())
            ))
            ahorro_id = cursor.lastrowid

        conn.commit()
        return ahorro_id


def obtener_ahorros(filtros: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Obtiene lista paginada de ahorros con filtros opcionales.

    Args:
        filtros: Diccionario con filtros (fecha_desde, fecha_hasta, categoria, tipo, page, per_page)

    Returns:
        Diccionario con items, total, page, per_page
    """
    filtros = filtros or {}
    page = filtros.get('page', 1)
    per_page = filtros.get('per_page', 20)
    offset = (page - 1) * per_page

    # Construir WHERE clause dinámicamente
    conditions = []
    params = {}

    if filtros.get('fecha_desde'):
        conditions.append("fecha_realizacion >= :fecha_desde")
        params['fecha_desde'] = filtros['fecha_desde']

    if filtros.get('fecha_hasta'):
        conditions.append("fecha_realizacion <= :fecha_hasta")
        params['fecha_hasta'] = filtros['fecha_hasta']

    if filtros.get('categoria'):
        conditions.append("categoria = :categoria")
        params['categoria'] = filtros['categoria']

    if filtros.get('tipo'):
        conditions.append("tipo = :tipo")
        params['tipo'] = filtros['tipo']

    if filtros.get('registrado_por'):
        conditions.append("registrado_por = :registrado_por")
        params['registrado_por'] = filtros['registrado_por']

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    conn = get_db_connection()
    cursor = conn.cursor()

    # Convertir params para SQLite si es necesario
    if not is_using_postgresql():
        # Convertir named params a positional
        query_count = f"SELECT COUNT(*) FROM ahorro_costo WHERE {where_clause}"
        query_data = f"""
            SELECT
                id, categoria, tipo, monto, moneda, material_codigo,
                proveedor_cuit, solicitud_id, descripcion, registrado_por,
                fecha_realizacion, created_at
            FROM ahorro_costo
            WHERE {where_clause}
            ORDER BY fecha_realizacion DESC
            LIMIT ? OFFSET ?
        """
        # Reemplazar :param por ?
        for key in params:
            query_count = query_count.replace(f":{key}", "?")
            query_data = query_data.replace(f":{key}", "?")

        cursor.execute(query_count, tuple(params.values()))
        total = cursor.fetchone()[0]

        cursor.execute(query_data, tuple(params.values()) + (per_page, offset))
    else:
        params['limit'] = per_page
        params['offset'] = offset

        cursor.execute(f"SELECT COUNT(*) FROM ahorro_costo WHERE {where_clause}", params)
        total = cursor.fetchone()[0]

        cursor.execute(f"""
            SELECT
                id, categoria, tipo, monto, moneda, material_codigo,
                proveedor_cuit, solicitud_id, descripcion, registrado_por,
                fecha_realizacion, created_at
            FROM ahorro_costo
            WHERE {where_clause}
            ORDER BY fecha_realizacion DESC
            LIMIT %(limit)s OFFSET %(offset)s
        """, params)

    columns = [desc[0] for desc in cursor.description]
    items = [dict(zip(columns, row)) for row in cursor.fetchall()]
    cursor.close()
    conn.close()

    return {
        'items': items,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page
    }


def obtener_kpis_ahorro() -> Dict[str, Any]:
    """
    Obtiene KPIs de ahorros (YTD, MTD, hard vs soft).

    Returns:
        Diccionario con totales y breakdowns
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    now = datetime.now()
    year_start = f"{now.year}-01-01"
    month_start = f"{now.year}-{now.month:02d}-01"

    # YTD total
    if is_using_postgresql():
        cursor.execute("""
            SELECT COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= %s
        """, (year_start,))
    else:
        cursor.execute("""
            SELECT COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= ?
        """, (year_start,))
    ytd_total = cursor.fetchone()[0]

    # MTD total
    if is_using_postgresql():
        cursor.execute("""
            SELECT COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= %s
        """, (month_start,))
    else:
        cursor.execute("""
            SELECT COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= ?
        """, (month_start,))
    mtd_total = cursor.fetchone()[0]

    # Hard vs Soft breakdown (YTD)
    if is_using_postgresql():
        cursor.execute("""
            SELECT tipo, COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= %s
            GROUP BY tipo
        """, (year_start,))
    else:
        cursor.execute("""
            SELECT tipo, COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= ?
            GROUP BY tipo
        """, (year_start,))

    tipo_breakdown = {row[0]: row[1] for row in cursor.fetchall()}

    # Breakdown por categoría (YTD)
    if is_using_postgresql():
        cursor.execute("""
            SELECT categoria, COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= %s
            GROUP BY categoria
        """, (year_start,))
    else:
        cursor.execute("""
            SELECT categoria, COALESCE(SUM(monto), 0) as total
            FROM ahorro_costo
            WHERE fecha_realizacion >= ?
            GROUP BY categoria
        """, (year_start,))

    categoria_breakdown = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.close()
    conn.close()

    return {
        'ytd_total': ytd_total,
        'mtd_total': mtd_total,
        'hard_savings': tipo_breakdown.get('hard', 0),
        'soft_savings': tipo_breakdown.get('soft', 0),
        'por_categoria': categoria_breakdown
    }


def crear_meta(data: Dict[str, Any], user_id: int) -> int:
    """
    Crea o actualiza una meta de ahorro.

    Args:
        data: Diccionario con categoria, periodo, meta_monto, buyer_id (opcional)
        user_id: ID del usuario que crea la meta

    Returns:
        ID de la meta
    """
    with get_db_transaction() as (conn, cursor):
        buyer_id = data.get('buyer_id')

        if is_using_postgresql():
            # Upsert en PostgreSQL
            cursor.execute("""
                INSERT INTO meta_ahorro (categoria, periodo, meta_monto, buyer_id)
                VALUES (%(categoria)s, %(periodo)s, %(meta_monto)s, %(buyer_id)s)
                ON CONFLICT (categoria, periodo, buyer_id)
                DO UPDATE SET
                    meta_monto = EXCLUDED.meta_monto,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """, {
                'categoria': data['categoria'],
                'periodo': data['periodo'],
                'meta_monto': data['meta_monto'],
                'buyer_id': buyer_id
            })
            meta_id = cursor.fetchone()[0]
        else:
            # Upsert en SQLite (INSERT OR REPLACE)
            cursor.execute("""
                INSERT OR REPLACE INTO meta_ahorro (categoria, periodo, meta_monto, buyer_id, updated_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (
                data['categoria'],
                data['periodo'],
                data['meta_monto'],
                buyer_id
            ))
            meta_id = cursor.lastrowid

        conn.commit()
        return meta_id


def obtener_metas(periodo: str) -> List[Dict[str, Any]]:
    """
    Obtiene metas de ahorro para un periodo.

    Args:
        periodo: Periodo en formato YYYYMM

    Returns:
        Lista de metas
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    if is_using_postgresql():
        cursor.execute("""
            SELECT id, categoria, periodo, meta_monto, buyer_id, created_at, updated_at
            FROM meta_ahorro
            WHERE periodo = %s
            ORDER BY categoria, buyer_id
        """, (periodo,))
    else:
        cursor.execute("""
            SELECT id, categoria, periodo, meta_monto, buyer_id, created_at, updated_at
            FROM meta_ahorro
            WHERE periodo = ?
            ORDER BY categoria, buyer_id
        """, (periodo,))

    columns = [desc[0] for desc in cursor.description]
    metas = [dict(zip(columns, row)) for row in cursor.fetchall()]
    cursor.close()
    conn.close()

    return metas


def obtener_progreso_metas(periodo: str) -> List[Dict[str, Any]]:
    """
    Obtiene progreso de metas vs ahorros reales para un periodo.

    Args:
        periodo: Periodo en formato YYYYMM (ej: 202602)

    Returns:
        Lista de metas con progreso
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Convertir periodo YYYYMM a rango de fechas
    year = int(periodo[:4])
    month = int(periodo[4:6])
    fecha_desde = f"{year}-{month:02d}-01"

    # Calcular último día del mes
    if month == 12:
        next_month = f"{year + 1}-01-01"
    else:
        next_month = f"{year}-{month + 1:02d}-01"

    if is_using_postgresql():
        cursor.execute("""
            SELECT
                m.id,
                m.categoria,
                m.periodo,
                m.meta_monto,
                m.buyer_id,
                COALESCE(SUM(a.monto), 0) as monto_actual
            FROM meta_ahorro m
            LEFT JOIN ahorro_costo a ON
                m.categoria = a.categoria
                AND a.fecha_realizacion >= %s
                AND a.fecha_realizacion < %s
                AND (m.buyer_id IS NULL OR m.buyer_id = a.registrado_por)
            WHERE m.periodo = %s
            GROUP BY m.id, m.categoria, m.periodo, m.meta_monto, m.buyer_id
            ORDER BY m.categoria, m.buyer_id
        """, (fecha_desde, next_month, periodo))
    else:
        cursor.execute("""
            SELECT
                m.id,
                m.categoria,
                m.periodo,
                m.meta_monto,
                m.buyer_id,
                COALESCE(SUM(a.monto), 0) as monto_actual
            FROM meta_ahorro m
            LEFT JOIN ahorro_costo a ON
                m.categoria = a.categoria
                AND a.fecha_realizacion >= ?
                AND a.fecha_realizacion < ?
                AND (m.buyer_id IS NULL OR m.buyer_id = a.registrado_por)
            WHERE m.periodo = ?
            GROUP BY m.id, m.categoria, m.periodo, m.meta_monto, m.buyer_id
            ORDER BY m.categoria, m.buyer_id
        """, (fecha_desde, next_month, periodo))

    columns = [desc[0] for desc in cursor.description]
    resultados = []
    for row in cursor.fetchall():
        item = dict(zip(columns, row))
        item['porcentaje_cumplimiento'] = (
            (item['monto_actual'] / item['meta_monto'] * 100)
            if item['meta_monto'] > 0 else 0
        )
        resultados.append(item)

    cursor.close()
    conn.close()

    return resultados


def exportar_ahorros_excel(filtros: Optional[Dict[str, Any]] = None) -> bytes:
    """
    Exporta ahorros a Excel.

    Args:
        filtros: Filtros a aplicar (mismos que obtener_ahorros)

    Returns:
        Bytes del archivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecute: pip install openpyxl")

    # Obtener todos los ahorros (sin paginación)
    filtros_export = filtros.copy() if filtros else {}
    filtros_export['per_page'] = 10000  # Límite alto para export
    data = obtener_ahorros(filtros_export)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ahorros"

    # Headers
    headers = [
        'ID', 'Categoría', 'Tipo', 'Monto', 'Moneda', 'Material',
        'Proveedor CUIT', 'Solicitud ID', 'Descripción',
        'Registrado Por', 'Fecha Realización', 'Fecha Creación'
    ]

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, item in enumerate(data['items'], 2):
        ws.cell(row=row_num, column=1, value=item['id'])
        ws.cell(row=row_num, column=2, value=item['categoria'])
        ws.cell(row=row_num, column=3, value=item['tipo'])
        ws.cell(row=row_num, column=4, value=item['monto'])
        ws.cell(row=row_num, column=5, value=item['moneda'])
        ws.cell(row=row_num, column=6, value=item.get('material_codigo', ''))
        ws.cell(row=row_num, column=7, value=item.get('proveedor_cuit', ''))
        ws.cell(row=row_num, column=8, value=item.get('solicitud_id', ''))
        ws.cell(row=row_num, column=9, value=item.get('descripcion', ''))
        ws.cell(row=row_num, column=10, value=item['registrado_por'])
        ws.cell(row=row_num, column=11, value=item['fecha_realizacion'])
        ws.cell(row=row_num, column=12, value=item.get('created_at', ''))

    # Auto-ajustar columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()
