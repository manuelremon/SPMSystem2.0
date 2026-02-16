"""
Dashboard routes - Sistema de dashboards editables tipo spreadsheet

Endpoints:
- GET    /api/dashboards                    - Listar dashboards del usuario
- POST   /api/dashboards                    - Crear dashboard
- GET    /api/dashboards/:uuid              - Obtener dashboard con sheets
- PUT    /api/dashboards/:uuid              - Actualizar dashboard
- DELETE /api/dashboards/:uuid              - Eliminar dashboard
- POST   /api/dashboards/:uuid/favorite     - Toggle favorito
- POST   /api/dashboards/:uuid/share        - Crear link compartido
- GET    /api/dashboards/:uuid/shares       - Listar links compartidos
- DELETE /api/dashboards/shares/:id         - Revocar link compartido
- GET    /api/shared/:token                 - Acceso publico por token
- PUT    /api/dashboards/sheets/:id         - Actualizar datos de hoja
- POST   /api/dashboards/:uuid/sheets       - Agregar hoja
- DELETE /api/dashboards/sheets/:id         - Eliminar hoja
- POST   /api/dashboards/:uuid/datasources  - Agregar fuente de datos
- DELETE /api/dashboards/datasources/:id    - Eliminar fuente de datos
- POST   /api/dashboard-functions/execute   - Ejecutar formulas SPM
- GET    /api/dashboard-functions/catalog   - Catalogo de formulas
- POST   /api/dashboards/:uuid/export       - Exportar a Excel
- GET    /api/dashboard-grupos              - Listar grupos
- POST   /api/dashboard-grupos              - Crear grupo
"""

import io
import logging

from flask import Blueprint, g, jsonify, request, send_file

logger = logging.getLogger(__name__)

from backend.core.dashboard_schemas import (
    SPM_FORMULAS_CATALOG,
    CreateDashboardRequest,
    CreateShareRequest,
    UpdateDashboardRequest,
)
from backend.core.roles import normalize_roles, require_auth
from backend.core.user_helpers import get_current_user_id
from backend.services.dashboard_service import (
    DashboardDataSourceService,
    DashboardGrupoService,
    DashboardPermisoService,
    DashboardService,
    DashboardShareService,
    DashboardSheetService,
)

bp = Blueprint("dashboards", __name__, url_prefix="/api")


def _get_user_id():
    """Obtiene el ID del usuario autenticado"""
    return get_current_user_id()


def _get_user_roles():
    """Obtiene los roles del usuario autenticado desde g.user"""
    user = getattr(g, "user", None)
    if not user:
        return []
    rol = user.get("rol", "") if isinstance(user, dict) else ""
    return [r.lower() for r in normalize_roles(rol)]


# ============================================================================
# Dashboard CRUD
# ============================================================================


@bp.route("/dashboards", methods=["GET"])
@require_auth
def list_dashboards():
    """Lista todos los dashboards accesibles por el usuario"""
    user_id = _get_user_id()
    grupo_id = request.args.get("grupo_id", type=int)
    include_shared = request.args.get("include_shared", "true").lower() == "true"

    dashboards = DashboardService.list_all(
        owner_id=user_id,
        user_roles=_get_user_roles(),
        grupo_id=grupo_id,
        include_shared=include_shared,
    )

    return jsonify({
        "success": True,
        "dashboards": [d.to_dict() for d in dashboards],
        "total": len(dashboards),
    })


@bp.route("/dashboards", methods=["POST"])
@require_auth
def create_dashboard():
    """Crea un nuevo dashboard"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    if not data.get("nombre"):
        return jsonify({"success": False, "error": "El nombre es requerido"}), 400

    req = CreateDashboardRequest.from_dict(data, user_id)
    result = DashboardService.create(req)

    if not result.success:
        return jsonify({
            "success": False,
            "error": result.error_message,
            "error_code": result.error_code,
        }), 400

    return jsonify(result.to_dict()), 201


@bp.route("/dashboards/<uuid>", methods=["GET"])
@require_auth
def get_dashboard(uuid):
    """Obtiene un dashboard con todas sus hojas y datasources"""
    user_id = _get_user_id()

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)

    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    # Obtener nivel de permiso del usuario
    perm = DashboardService.get_permission_level(dashboard.id, user_id, dashboard.owner_id)

    return jsonify({
        "success": True,
        "dashboard": dashboard.to_dict(include_sheets=True, include_datasources=True),
        "permiso": perm,
    })


@bp.route("/dashboards/<uuid>", methods=["PUT"])
@require_auth
def update_dashboard(uuid):
    """Actualiza un dashboard existente"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    req = UpdateDashboardRequest.from_dict(data)
    result = DashboardService.update(uuid, req, user_id)

    if not result.success:
        status = 404 if result.error_code == "NOT_FOUND" else 403 if result.error_code == "FORBIDDEN" else 400
        return jsonify({
            "success": False,
            "error": result.error_message,
            "error_code": result.error_code,
        }), status

    return jsonify(result.to_dict())


@bp.route("/dashboards/<uuid>", methods=["DELETE"])
@require_auth
def delete_dashboard(uuid):
    """Elimina un dashboard"""
    user_id = _get_user_id()

    if DashboardService.delete(uuid, user_id):
        return jsonify({"success": True, "message": "Dashboard eliminado"})

    return jsonify({"success": False, "error": "No se pudo eliminar el dashboard"}), 400


@bp.route("/dashboards/<uuid>/favorite", methods=["POST"])
@require_auth
def toggle_favorite(uuid):
    """Alterna el estado de favorito"""
    user_id = _get_user_id()

    result = DashboardService.toggle_favorite(uuid, user_id)

    if result is None:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    return jsonify({"success": True, "es_favorito": result})


# ============================================================================
# Comparticion
# ============================================================================


@bp.route("/dashboards/<uuid>/share", methods=["POST"])
@require_auth
def create_share(uuid):
    """Crea un link de comparticion"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    # Obtener dashboard para obtener su ID
    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    req = CreateShareRequest.from_dict(data, dashboard.id, user_id)
    share = DashboardShareService.create_share(req)

    if not share:
        return jsonify({"success": False, "error": "No se pudo crear el link"}), 400

    return jsonify({
        "success": True,
        "share": share.to_dict(),
        "url": f"/shared/{share.token}",
    }), 201


@bp.route("/dashboards/<uuid>/shares", methods=["GET"])
@require_auth
def list_shares(uuid):
    """Lista los links de comparticion de un dashboard"""
    user_id = _get_user_id()

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    shares = DashboardShareService.list_shares(dashboard.id, user_id)

    return jsonify({
        "success": True,
        "shares": [s.to_dict() for s in shares],
    })


@bp.route("/dashboards/shares/<int:share_id>", methods=["DELETE"])
@require_auth
def revoke_share(share_id):
    """Revoca un link de comparticion"""
    user_id = _get_user_id()

    if DashboardShareService.revoke_share(share_id, user_id):
        return jsonify({"success": True, "message": "Link revocado"})

    return jsonify({"success": False, "error": "No se pudo revocar el link"}), 400


@bp.route("/shared/<token>", methods=["GET", "POST"])
def access_shared(token):
    """Acceso publico a dashboard compartido"""
    password = None
    if request.method == "POST":
        data = request.get_json() or {}
        password = data.get("password")

    result = DashboardShareService.get_by_token(token, password)

    if not result:
        return jsonify({"success": False, "error": "Link no encontrado"}), 404

    if "error" in result:
        status = 401 if result["error"] == "PASSWORD_REQUIRED" else 403
        return jsonify({"success": False, **result}), status

    return jsonify({
        "success": True,
        **result,
    })


# ============================================================================
# Hojas de Calculo
# ============================================================================


@bp.route("/dashboards/sheets/<int:sheet_id>", methods=["PUT"])
@require_auth
def update_sheet(sheet_id):
    """Actualiza los datos de una hoja"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    sheet_data = data.get("data")
    if not sheet_data:
        return jsonify({"success": False, "error": "Los datos son requeridos"}), 400

    if DashboardSheetService.update_data(sheet_id, sheet_data, user_id):
        return jsonify({"success": True, "message": "Hoja actualizada"})

    return jsonify({"success": False, "error": "No se pudo actualizar la hoja"}), 400


@bp.route("/dashboards/<uuid>/sheets", methods=["POST"])
@require_auth
def add_sheet(uuid):
    """Agrega una nueva hoja al dashboard"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    nombre = data.get("nombre", f"Hoja {len(dashboard.sheets) + 1}")
    sheet = DashboardSheetService.add_sheet(dashboard.id, nombre, user_id)

    if not sheet:
        return jsonify({"success": False, "error": "No se pudo agregar la hoja"}), 400

    return jsonify({
        "success": True,
        "sheet": sheet.to_dict(),
    }), 201


@bp.route("/dashboards/sheets/<int:sheet_id>", methods=["DELETE"])
@require_auth
def delete_sheet(sheet_id):
    """Elimina una hoja (debe quedar al menos una)"""
    user_id = _get_user_id()

    if DashboardSheetService.delete_sheet(sheet_id, user_id):
        return jsonify({"success": True, "message": "Hoja eliminada"})

    return jsonify({"success": False, "error": "No se pudo eliminar la hoja"}), 400


# ============================================================================
# Fuentes de Datos
# ============================================================================


@bp.route("/dashboards/<uuid>/datasources", methods=["POST"])
@require_auth
def add_datasource(uuid):
    """Agrega una fuente de datos"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    nombre = data.get("nombre")
    tipo = data.get("tipo")

    if not nombre or not tipo:
        return jsonify({"success": False, "error": "Nombre y tipo son requeridos"}), 400

    ds = DashboardDataSourceService.create(
        dashboard_id=dashboard.id,
        nombre=nombre,
        tipo=tipo,
        user_id=user_id,
        query=data.get("query"),
        filtros=data.get("filtros"),
        config=data.get("config"),
    )

    if not ds:
        return jsonify({"success": False, "error": "No se pudo agregar la fuente de datos"}), 400

    return jsonify({
        "success": True,
        "datasource": ds.to_dict(),
    }), 201


@bp.route("/dashboards/datasources/<int:datasource_id>", methods=["DELETE"])
@require_auth
def delete_datasource(datasource_id):
    """Elimina una fuente de datos"""
    user_id = _get_user_id()

    if DashboardDataSourceService.delete(datasource_id, user_id):
        return jsonify({"success": True, "message": "Fuente de datos eliminada"})

    return jsonify({"success": False, "error": "No se pudo eliminar la fuente de datos"}), 400


# ============================================================================
# Formulas SPM
# ============================================================================


@bp.route("/dashboard-functions/catalog", methods=["GET"])
@require_auth
def get_formula_catalog():
    """Retorna el catalogo de formulas SPM disponibles"""
    # Convertir enum a string para JSON
    catalog = {}
    for name, info in SPM_FORMULAS_CATALOG.items():
        catalog[name] = {
            **info,
            "categoria": info["categoria"].value if hasattr(info["categoria"], "value") else info["categoria"],
        }

    return jsonify({
        "success": True,
        "formulas": catalog,
    })


@bp.route("/dashboard-functions/execute", methods=["POST"])
@require_auth
def execute_formula():
    """Ejecuta una formula SPM"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    formula = data.get("formula")
    params = data.get("params", [])

    if not formula:
        return jsonify({"success": False, "error": "Formula es requerida"}), 400

    from backend.services.dashboard_formulas import FormulaExecutor

    result = FormulaExecutor.execute(formula, params, user_id, _get_user_roles())
    return jsonify(result.to_dict())


# ============================================================================
# Exportacion
# ============================================================================


@bp.route("/dashboards/<uuid>/export", methods=["POST"])
@require_auth
def export_dashboard(uuid):
    """Exporta el dashboard a Excel"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    data.get("formato", "xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Border, Font, PatternFill, Side  # noqa: F401

        wb = Workbook()

        for idx, sheet in enumerate(dashboard.sheets):
            if idx == 0:
                ws = wb.active
                ws.title = sheet.nombre[:31]  # Excel limita a 31 caracteres
            else:
                ws = wb.create_sheet(title=sheet.nombre[:31])

            # Procesar celldata de Fortune Sheet
            celldata = sheet.data.get("celldata", [])
            for cell in celldata:
                row = cell.get("r", 0) + 1
                col = cell.get("c", 0) + 1
                v = cell.get("v", {})

                # Obtener valor
                value = v.get("v") if isinstance(v, dict) else v
                if value is not None:
                    ws.cell(row=row, column=col, value=value)

                    # Aplicar estilos basicos
                    if isinstance(v, dict):
                        cell_obj = ws.cell(row=row, column=col)
                        if v.get("bl"):  # Bold
                            cell_obj.font = Font(bold=True)
                        if v.get("bg"):  # Background color
                            cell_obj.fill = PatternFill(start_color=v["bg"].replace("#", ""), fill_type="solid")

        # Guardar en buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{dashboard.nombre.replace(' ', '_')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except ImportError:
        return jsonify({
            "success": False,
            "error": "openpyxl no instalado",
        }), 500
    except Exception as e:
        logger.error(f"Error exportando dashboard: {e}")
        return jsonify({
            "success": False,
            "error": "Error al exportar dashboard",
        }), 500


# ============================================================================
# Grupos
# ============================================================================


@bp.route("/dashboard-grupos", methods=["GET"])
@require_auth
def list_grupos():
    """Lista todos los grupos accesibles"""
    user_id = _get_user_id()
    include_public = request.args.get("include_public", "true").lower() == "true"

    grupos = DashboardGrupoService.list_all(user_id, include_public)

    return jsonify({
        "success": True,
        "grupos": [g.to_dict() for g in grupos],
    })


@bp.route("/dashboard-grupos", methods=["POST"])
@require_auth
def create_grupo():
    """Crea un nuevo grupo"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    nombre = data.get("nombre")
    if not nombre:
        return jsonify({"success": False, "error": "El nombre es requerido"}), 400

    grupo = DashboardGrupoService.create(
        nombre=nombre,
        owner_id=user_id,
        descripcion=data.get("descripcion"),
        icono=data.get("icono", "folder"),
        color=data.get("color", "#6366f1"),
    )

    return jsonify({
        "success": True,
        "grupo": grupo.to_dict(),
    }), 201


@bp.route("/dashboard-grupos/<int:grupo_id>", methods=["DELETE"])
@require_auth
def delete_grupo(grupo_id):
    """Elimina un grupo"""
    user_id = _get_user_id()

    if DashboardGrupoService.delete(grupo_id, user_id):
        return jsonify({"success": True, "message": "Grupo eliminado"})

    return jsonify({"success": False, "error": "No se pudo eliminar el grupo"}), 400


# ============================================================================
# Permisos
# ============================================================================


@bp.route("/dashboards/<uuid>/permisos", methods=["GET"])
@require_auth
def list_permisos(uuid):
    """Lista los permisos de un dashboard"""
    user_id = _get_user_id()

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    # Solo el owner puede ver permisos
    if dashboard.owner_id != user_id:
        return jsonify({"success": False, "error": "Sin permiso"}), 403

    permisos = DashboardPermisoService.list_permisos(dashboard.id)

    return jsonify({
        "success": True,
        "permisos": [p.to_dict() for p in permisos],
    })


@bp.route("/dashboards/<uuid>/permisos", methods=["POST"])
@require_auth
def grant_permiso(uuid):
    """Otorga permiso a un usuario"""
    user_id = _get_user_id()
    data = request.get_json() or {}

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    usuario_id = data.get("usuario_id")
    permiso = data.get("permiso", "view")

    if not usuario_id:
        return jsonify({"success": False, "error": "usuario_id es requerido"}), 400

    result = DashboardPermisoService.grant(dashboard.id, usuario_id, permiso, user_id)

    if not result:
        return jsonify({"success": False, "error": "No se pudo otorgar el permiso"}), 400

    return jsonify({
        "success": True,
        "permiso": result.to_dict(),
    }), 201


@bp.route("/dashboards/<uuid>/permisos/<usuario_id>", methods=["DELETE"])
@require_auth
def revoke_permiso(uuid, usuario_id):
    """Revoca el permiso de un usuario"""
    user_id = _get_user_id()

    dashboard = DashboardService.get_by_uuid(uuid, user_id, check_permission=True)
    if not dashboard:
        return jsonify({"success": False, "error": "Dashboard no encontrado"}), 404

    if DashboardPermisoService.revoke(dashboard.id, usuario_id, user_id):
        return jsonify({"success": True, "message": "Permiso revocado"})

    return jsonify({"success": False, "error": "No se pudo revocar el permiso"}), 400


# ============================================================================
# Resumen Ejecutivo
# ============================================================================


@bp.route("/dashboards/resumen-ejecutivo", methods=["GET"])
@require_auth
def resumen_ejecutivo():
    """
    Endpoint consolidado que devuelve un resumen ejecutivo del estado del día.
    Combina datos de solicitudes, presupuesto, alertas MRP y SLA en una sola llamada.

    Returns:
        {solicitudes_hoy, presupuesto_resumen, alertas_mrp_criticas, sla_breaches}
    """
    from datetime import datetime as _dt

    from backend.core.db import get_db_connection, is_using_postgresql, sql_date_relative

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # 1. Solicitudes del día / pendientes
            if is_using_postgresql():
                cursor.execute("""
                    SELECT
                        COUNT(*) FILTER (WHERE DATE(created_at) = CURRENT_DATE) as hoy,
                        COUNT(*) FILTER (WHERE status = 'submitted') as pendientes_aprobacion,
                        COUNT(*) FILTER (WHERE status = 'approved') as pendientes_planificacion,
                        COUNT(*) FILTER (WHERE status IN ('processing', 'dispatched')) as en_proceso,
                        COUNT(*) as total
                    FROM solicitud
                """)
            else:
                cursor.execute("""
                    SELECT
                        SUM(CASE WHEN DATE(created_at) = DATE('now') THEN 1 ELSE 0 END) as hoy,
                        SUM(CASE WHEN status = 'submitted' THEN 1 ELSE 0 END) as pendientes_aprobacion,
                        SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as pendientes_planificacion,
                        SUM(CASE WHEN status IN ('processing', 'dispatched') THEN 1 ELSE 0 END) as en_proceso,
                        COUNT(*) as total
                    FROM solicitud
                """)
            row = cursor.fetchone()
            sol = dict(row) if row else {}

            # 2. Presupuesto resumen
            cursor.execute("""
                SELECT
                    COALESCE(SUM(monto_usd), 0) as total,
                    COALESCE(SUM(saldo_usd), 0) as disponible,
                    COALESCE(SUM(monto_usd) - SUM(saldo_usd), 0) as utilizado
                FROM presupuesto
                WHERE monto_usd > 0
            """)
            prow = cursor.fetchone()
            presupuesto = dict(prow) if prow else {}
            total_pres = float(presupuesto.get('total') or 0)
            presupuesto['porcentaje_usado'] = round(
                (float(presupuesto.get('utilizado') or 0) / total_pres * 100) if total_pres > 0 else 0, 1
            )

            # 3. Alertas MRP críticas (materiales con stock 0 o bajo punto de pedido)
            alertas_mrp = {"criticas": 0, "altas": 0}
            try:
                db_name = "sap_data"
                with get_db_connection(db_name) as conn2:
                    cur2 = conn2.cursor()
                    cur2.execute("""
                        SELECT
                            SUM(CASE WHEN s.stock_actual <= 0 THEN 1 ELSE 0 END) as criticas,
                            SUM(CASE WHEN s.stock_actual > 0 AND s.stock_actual < m.punto_de_pedido THEN 1 ELSE 0 END) as altas
                        FROM materiales_bbdd m
                        LEFT JOIN (
                            SELECT material, centro, SUM(stock) as stock_actual
                            FROM stock GROUP BY material, centro
                        ) s ON m.codigo_material = s.material AND m.centro = s.centro
                        WHERE m.punto_de_pedido > 0
                    """)
                    arow = cur2.fetchone()
                    if arow:
                        alertas_mrp = {
                            "criticas": int(dict(arow).get("criticas") or 0),
                            "altas": int(dict(arow).get("altas") or 0)
                        }
            except Exception as e:
                logger.warning(f"Error obteniendo alertas MRP: {e}")

            # 4. SLA breaches (solicitudes que exceden tiempo límite)
            sla_breaches = 0
            try:
                cursor.execute(f"""
                    SELECT COUNT(*) as total
                    FROM solicitud
                    WHERE status = 'submitted'
                    AND created_at < {sql_date_relative(days=-3)}
                """)
                sla_row = cursor.fetchone()
                sla_breaches = int((dict(sla_row) if sla_row else {}).get('total', 0))
            except Exception as e:
                logger.warning(f"Error obteniendo SLA breaches: {e}")

        return jsonify({
            "ok": True,
            "data": {
                "solicitudes": {
                    "hoy": int(sol.get("hoy") or 0),
                    "pendientes_aprobacion": int(sol.get("pendientes_aprobacion") or 0),
                    "pendientes_planificacion": int(sol.get("pendientes_planificacion") or 0),
                    "en_proceso": int(sol.get("en_proceso") or 0),
                    "total": int(sol.get("total") or 0)
                },
                "presupuesto": {
                    "total": float(presupuesto.get("total") or 0),
                    "disponible": float(presupuesto.get("disponible") or 0),
                    "utilizado": float(presupuesto.get("utilizado") or 0),
                    "porcentaje_usado": presupuesto.get("porcentaje_usado", 0)
                },
                "alertas_mrp": alertas_mrp,
                "sla_breaches": sla_breaches,
                "generado_at": _dt.now().isoformat()
            }
        })

    except Exception as e:
        logger.error(f"Error en resumen ejecutivo: {e}")
        return jsonify({"ok": False, "error": {"code": "resumen_error", "message": "Error al generar resumen ejecutivo"}}), 500
